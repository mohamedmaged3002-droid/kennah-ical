#!/usr/bin/env python3
"""out/changes.json -> out/kennah-price-changes.xlsx (changed units only)."""
import json, pathlib
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = pathlib.Path(__file__).parent
src = ROOT / "out" / "changes.json"
if not src.exists():
    print("no changes.json — nothing to build"); raise SystemExit(0)
d = json.loads(src.read_text())
if not any(d.get(k) for k in ("changed", "added", "removed", "newOnSite", "goneFromSite")):
    print("no changes — no sheet"); raise SystemExit(0)

wb = Workbook(); ws = wb.active; ws.title = "Price changes"
hdr = ["wp_post_id", "unit", "date", "old USD", "new USD", "change USD", "direction"]
ws.append(hdr)
for c in range(1, len(hdr) + 1):
    ws.cell(1, c).font = Font(bold=True, color="FFFFFF")
    ws.cell(1, c).fill = PatternFill("solid", fgColor="1F4E79")

up = PatternFill("solid", fgColor="FCE4D6")
down = PatternFill("solid", fgColor="E2EFDA")
for u in d.get("changed", []):
    for n in u["nights"]:
        delta = n["to"] - n["from"]
        ws.append([u["wp"], u["name"], n["date"], n["from"], n["to"], delta,
                   "increase" if delta > 0 else "decrease"])
        for c in range(1, len(hdr) + 1):
            ws.cell(ws.max_row, c).fill = up if delta > 0 else down

for i, w in enumerate([12, 44, 12, 10, 10, 12, 11], start=1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(hdr))}{ws.max_row}"

if any(d.get(k) for k in ("added", "removed", "newOnSite", "goneFromSite")):
    r = wb.create_sheet("Roster")
    r.append(["change", "wp_post_id", "unit", "slug"])
    for c in range(1, 5):
        r.cell(1, c).font = Font(bold=True, color="FFFFFF")
        r.cell(1, c).fill = PatternFill("solid", fgColor="1F4E79")
    for n in d.get("newOnSite", []):
        r.append(["NEW listing — needs onboarding", "", n.get("name", ""), n.get("slug", "")])
    for g in d.get("goneFromSite", []):
        r.append(["GONE from site — check delist", g.get("wp", ""), g.get("name", ""), g.get("slug", "")])
    for wp in d.get("added", []): r.append(["now priced", wp, "", ""])
    for wp in d.get("removed", []): r.append(["lost prices — check calendar", wp, "", ""])
    for i, w in enumerate([32, 13, 42, 42], start=1):
        r.column_dimensions[chr(64 + i)].width = w

out = ROOT / "out" / "kennah-price-changes.xlsx"
wb.save(out)
print(f"wrote {out} ({ws.max_row - 1} changed nights)")
